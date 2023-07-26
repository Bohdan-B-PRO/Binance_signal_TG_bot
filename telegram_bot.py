import os
from binance.client import Client
from datetime import datetime
import time
from openpyxl import Workbook
import telebot

# Ключи для Бинанс
api_key = ('nSsyw7ejwwTKpQSq5rLeOvC2ShilJUZXlKHBCKoSbqpnvHbl3CLoPa6FC0PKh2k8')
api_secret = ('tnUrNDk4HEseTMEau5xiQUzueWlz8e9l8Dtct21rnS1WkhtTQ1Pf173EtVoO1QHy')
client = Client(api_key, api_secret)

# Создаю таблицу для сбора данных
wb = Workbook()
ws = wb.active
# Расписываю названия колонок для данных
ws['A1'] = 'Average'
ws['B1'] = 'Futures'
ws['C1'] = 'Percentage'
ws['D1'] = 'Time'

# Задаю торговую пару:
ASSET = 'BTCUSDT'


def price(symbol):
    try:
        price = client.get_avg_price(symbol=symbol, requests_params={"timeout": 2})['price']
        return float(price)
    except Exception as e:
        print(e)


def priceF(symbol):
    try:
        priceF = client.futures_symbol_ticker(symbol=symbol, requests_params={"timeout": 2})["price"]
        return float(priceF)
    except Exception as e:
        print(e)


bot = telebot.TeleBot("6503824319:AAEs6rBvEc9-oYP07UPF-GbBaAWT0-IhcFk")

# ID или список ID получателей
ID = 379834541

# Функция отправки сообщений
def message(text):
    bot.send_message(ID, text)


def message_signal():
    message(f"Сигнал по торговой паре: {ASSET}")

TIME = 1

# Процент выше которого я начинаю получать сигналы
GROWTH_PERCENT = 0.053

while True:
    # Сравниваю цену фьючерсов со средней ценой
    FIRST_PRICE = price(ASSET)
    PRICEF = priceF(ASSET)
    PERCENT = ((PRICEF - FIRST_PRICE) / FIRST_PRICE) * 100

    if PERCENT >= GROWTH_PERCENT:
        message_signal()
        print(ASSET)
        print("LOOK")
        print(PERCENT)
        times = datetime.now().strftime("%H:%M:%S")
        print(times)
        ws.append([FIRST_PRICE, PRICEF, PERCENT, times])
        wb.save("data.xlsx")

    time.sleep(TIME)











